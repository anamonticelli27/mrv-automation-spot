import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment

from frete_spot import config
from frete_spot.excel.styles import (
    CENTER,
    FONT_BLACK,
    FONT_BLACK_BOLD,
    FONT_GREEN_BOLD,
    FONT_WHITE_BOLD,
    GREEN_DARK,
    GREEN_LIGHT,
    GREY_LIGHT,
    RED_LIGHT,
    WHITE_FILL,
    apply,
    thin_border,
)
from frete_spot.excel.workbook import (
    atualizar_chamado_vencedor,
    carregar_chamado,
    carregar_cotacoes,
)


def parse_valor(cotacao):
    try:
        return float(
            str(cotacao.get("Valor Frete (R$)", "0"))
            .replace("R$", "")
            .replace(".", "")
            .replace(",", ".")
            .strip()
        )
    except Exception:
        return float("inf")


def _filter_respondidas(cotacoes):
    respondidas = []
    for cotacao in cotacoes:
        val = cotacao.get("Valor Frete (R$)", "")
        if val != "" and val is not None:
            try:
                float(str(val).replace("R$", "").replace(".", "").replace(",", ".").strip())
                respondidas.append(cotacao)
            except Exception:
                pass
    return respondidas


def _build_mapa_workbook(chamado_num, chamado, respondidas, vencedor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa de Cotação"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 4

    num_fornecedores = len(respondidas)
    col_start = 6
    cols_per_supplier = 2

    for i in range(num_fornecedores):
        base = col_start + i * cols_per_supplier
        ws.column_dimensions[ws.cell(row=1, column=base).column_letter].width = 14
        ws.column_dimensions[ws.cell(row=1, column=base + 1).column_letter].width = 14

    for i, cot in enumerate(respondidas):
        base = col_start + i * cols_per_supplier
        label = f"FORNECEDOR {str(i + 1).zfill(2)}"
        ws.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base + 1)
        cell = ws.cell(row=4, column=base)
        cell.value = label
        cell.fill = GREEN_DARK
        cell.font = FONT_WHITE_BOLD
        cell.alignment = CENTER
        cell.border = thin_border()

    labels = [
        "CÓDIGO", "FORNECEDOR", "TELEFONE", "E-MAIL",
        "CONTATO", "PRAZO ENTREGA", "PRAZO DE PAGAMENTO", "OBSERVAÇÃO",
    ]
    fields = [
        "Número Chamado", "Transportadora", "Telefone",
        "Email", "Contato", "Prazo Entrega", "_pagamento", "Observações",
    ]

    for idx, (label, field) in enumerate(zip(labels, fields)):
        r = 5 + idx
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        apply(ws, r, 4, label, GREEN_DARK, FONT_WHITE_BOLD, CENTER)

        for i, cot in enumerate(respondidas):
            base = col_start + i * cols_per_supplier
            ws.merge_cells(start_row=r, start_column=base, end_row=r, end_column=base + 1)
            cell = ws.cell(row=r, column=base)

            if field == "_pagamento":
                val = "30 Dias"
            elif field == "Número Chamado":
                val = chamado_num
            else:
                val = cot.get(field, "")

            cell.value = val
            cell.fill = WHITE_FILL
            cell.font = FONT_BLACK
            cell.alignment = CENTER
            cell.border = thin_border()

    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=3)
    apply(ws, 13, 2, "DESCRIÇÃO", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)
    apply(ws, 13, 4, "QTD", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)
    ws.merge_cells(start_row=13, start_column=4, end_row=13, end_column=5)

    for i in range(num_fornecedores):
        base = col_start + i * cols_per_supplier
        apply(ws, 13, base, "VALOR\nUNITARIO", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)
        apply(ws, 13, base + 1, "VALOR TOTAL", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)

    assunto = chamado.get("Assunto", "Frete")
    tipo_mat = chamado.get("Tipo de Material", "")
    descricao = f"{assunto}" + (f" - {tipo_mat}" if tipo_mat else "")

    ws.merge_cells(start_row=14, start_column=2, end_row=14, end_column=3)
    apply(ws, 14, 2, descricao, WHITE_FILL, FONT_BLACK, CENTER)

    ws.merge_cells(start_row=14, start_column=4, end_row=14, end_column=5)
    apply(ws, 14, 4, 1, WHITE_FILL, FONT_BLACK, CENTER)

    for i, cot in enumerate(respondidas):
        base = col_start + i * cols_per_supplier
        val = parse_valor(cot)
        is_winner = cot == vencedor
        fill = GREEN_LIGHT if is_winner else RED_LIGHT

        cell_u = ws.cell(row=14, column=base)
        cell_u.value = val
        cell_u.number_format = "R$ #,##0.00"
        cell_u.fill = fill
        cell_u.font = FONT_BLACK_BOLD
        cell_u.alignment = CENTER
        cell_u.border = thin_border()

        cell_t = ws.cell(row=14, column=base + 1)
        cell_t.value = val
        cell_t.number_format = "R$ #,##0.00"
        cell_t.fill = fill
        cell_t.font = FONT_BLACK_BOLD
        cell_t.alignment = CENTER
        cell_t.border = thin_border()

    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=5)
    apply(ws, 15, 2, "VALOR TOTAL", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)

    for i, cot in enumerate(respondidas):
        base = col_start + i * cols_per_supplier
        val = parse_valor(cot)
        is_winner = cot == vencedor
        fill = GREEN_LIGHT if is_winner else RED_LIGHT

        ws.merge_cells(start_row=15, start_column=base, end_row=15, end_column=base + 1)
        cell = ws.cell(row=15, column=base)
        cell.value = val
        cell.number_format = "R$ #,##0.00"
        cell.fill = fill
        cell.font = FONT_BLACK_BOLD
        cell.alignment = CENTER
        cell.border = thin_border()

    ws.merge_cells(start_row=6, start_column=2, end_row=10, end_column=3)
    title_cell = ws.cell(row=6, column=2)
    title_cell.value = f"#{chamado_num}\n{assunto}"
    title_cell.font = FONT_GREEN_BOLD
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border = thin_border()

    return wb


def gerar_mapa(chamado_num):
    os.makedirs(config.MAPAS_DIR, exist_ok=True)

    chamado, wb_main = carregar_chamado(chamado_num)
    if not chamado:
        print(f"❌ Chamado {chamado_num} não encontrado na planilha.")
        return
    wb_main.close()

    cotacoes = carregar_cotacoes(chamado_num)
    respondidas = _filter_respondidas(cotacoes)

    if not respondidas:
        print("⚠️  Nenhuma cotação com valor preenchido encontrada.")
        print("    Preencha a coluna 'Valor Frete (R$)' na aba Cotações e tente novamente.")
        return

    respondidas.sort(key=parse_valor)
    vencedor = respondidas[0]

    wb = _build_mapa_workbook(chamado_num, chamado, respondidas, vencedor)

    data_hoje = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"Mapa_{chamado_num}_{data_hoje}.xlsx"
    caminho = os.path.join(config.MAPAS_DIR, nome_arquivo)
    wb.save(caminho)
    print(f"\n✅ Mapa salvo em: 3_Mapas_Cotacao/{nome_arquivo}")

    atualizar_chamado_vencedor(
        chamado_num,
        vencedor.get("Transportadora", ""),
        data_hoje,
    )

    print(f"\n🏆 Vencedor  : {vencedor.get('Transportadora')} — R$ {parse_valor(vencedor):,.2f}")
    print(f"📅 Data Mapa : {data_hoje}")
    print("\n📊 Resumo das cotações:")
    for i, c in enumerate(respondidas):
        tag = " 🏆 VENCEDOR" if c == vencedor else ""
        print(f"   {i + 1}. {c.get('Transportadora'):25s} R$ {parse_valor(c):>10,.2f}{tag}")


def main():
    print("=" * 50)
    print("   GERADOR DE MAPA DE COTAÇÃO")
    print("=" * 50)
    chamado_num = input("\n👉 Digite o número do chamado: ").strip()

    if not chamado_num:
        print("❌ Número do chamado não informado.")
        return

    gerar_mapa(chamado_num)
