from datetime import datetime

from openpyxl import load_workbook

from frete_spot import config
from frete_spot.config import (
    CHAMADOS_LEFT_ALIGN,
    COL_WIDTHS_CHAMADOS,
    COL_WIDTHS_COTACOES,
    COTACOES_LEFT_ALIGN,
    SLA_HEADER_NAMES,
)
from frete_spot.excel.styles import (
    CENTER,
    FONT_BOLD,
    FONT_NORMAL,
    FONT_SLA_GREEN,
    FONT_SLA_RED,
    FONT_SLA_YELLOW,
    GREEN_SLA,
    GREY_ROW,
    LEFT,
    RED_SLA,
    WHITE,
    YELLOW_EMPTY,
    YELLOW_SLA,
    set_column_widths,
    style_header_row,
    thin_border,
)


def _apply_sla_style(cell, today):
    try:
        sla_date = datetime.strptime(str(cell.value), "%d/%m/%Y").date()
        diff = (sla_date - today).days
        if diff < 0:
            cell.fill = RED_SLA
            cell.font = FONT_SLA_RED
        elif diff <= 2:
            cell.fill = YELLOW_SLA
            cell.font = FONT_SLA_YELLOW
        else:
            cell.fill = GREEN_SLA
            cell.font = FONT_SLA_GREEN
    except Exception:
        pass


def _format_chamados_tab(ws, today):
    headers = [c.value for c in ws[1]]
    set_column_widths(ws, headers, COL_WIDTHS_CHAMADOS)

    sla_cols = [headers.index(name) + 1 for name in SLA_HEADER_NAMES if name in headers]
    status_col = headers.index("Status Geral") + 1 if "Status Geral" in headers else None
    border = thin_border()

    style_header_row(ws, headers, border)
    ws.freeze_panes = "A2"

    for row_idx in range(2, ws.max_row + 1):
        row_fill = WHITE if row_idx % 2 == 0 else GREY_ROW
        ws.row_dimensions[row_idx].height = 20

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = FONT_NORMAL
            cell.alignment = LEFT if header in CHAMADOS_LEFT_ALIGN else CENTER
            cell.border = border

            if col_idx in sla_cols and cell.value:
                _apply_sla_style(cell, today)

            if status_col and col_idx == status_col:
                val = str(cell.value or "")
                if "Vencido" in val:
                    cell.fill = RED_SLA
                    cell.font = FONT_SLA_RED
                elif "prazo" in val.lower():
                    cell.fill = GREEN_SLA
                    cell.font = FONT_SLA_GREEN

            if header == "Transportadora Vencedora" and cell.value:
                cell.font = FONT_BOLD


def _format_cotacoes_tab(ws):
    headers = [c.value for c in ws[1]]
    set_column_widths(ws, headers, COL_WIDTHS_COTACOES)

    valor_col = headers.index("Valor Frete (R$)") + 1 if "Valor Frete (R$)" in headers else None
    data_col = headers.index("Data Resposta") + 1 if "Data Resposta" in headers else None
    border = thin_border()

    style_header_row(ws, headers, border)
    ws.freeze_panes = "A2"

    for row_idx in range(2, ws.max_row + 1):
        row_fill = WHITE if row_idx % 2 == 0 else GREY_ROW
        ws.row_dimensions[row_idx].height = 20

        data_val = ws.cell(row=row_idx, column=data_col).value if data_col else None
        valor_val = ws.cell(row=row_idx, column=valor_col).value if valor_col else None
        sem_resposta = not data_val and not valor_val

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = YELLOW_EMPTY if sem_resposta else row_fill
            cell.font = FONT_NORMAL
            cell.alignment = LEFT if header in COTACOES_LEFT_ALIGN else CENTER
            cell.border = border

            if col_idx == valor_col and cell.value:
                cell.font = FONT_BOLD
                cell.number_format = "R$ #,##0.00"


def formatar_planilha():
    wb = load_workbook(config.PLANILHA_PATH)
    today = datetime.now().date()

    _format_chamados_tab(wb["Chamados"], today)
    _format_cotacoes_tab(wb["Cotações"])

    wb.save(config.PLANILHA_PATH)
    wb.close()
    print("🎨 Planilha formatada com sucesso!")
