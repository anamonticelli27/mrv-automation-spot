import os
from datetime import datetime

import openpyxl
from openpyxl import load_workbook

from frete_spot import config
from frete_spot.config import HEADERS_CHAMADOS, HEADERS_COTACOES


def fmt_data(dt):
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y")
    return ""


def chamado_to_row(dados, status):
    sla_map = {
        "SLA Step 1 (Validar)": fmt_data(dados.get("SLA Step 1")),
        "SLA Step 2 (Cotar)": fmt_data(dados.get("SLA Step 2")),
        "SLA Step 3 (OK Cliente)": fmt_data(dados.get("SLA Step 3")),
        "SLA Step 4 (Programar)": fmt_data(dados.get("SLA Step 4")),
    }
    field_map = {
        "Número Chamado": dados.get("Número Chamado", ""),
        "Assunto": dados.get("Assunto", ""),
        "PEP / Diagrama de Rede": dados.get("PEP", ""),
        "Data Recebimento": fmt_data(dados.get("Data Recebimento")),
        "Categoria": dados.get("Categoria", ""),
        "Status Geral": status,
        "Subcategoria": dados.get("Subcategoria", ""),
        "Solicitante": dados.get("Solicitante", ""),
        "Email Solicitante": dados.get("Email Solicitante", ""),
        "Telefone": dados.get("Telefone", ""),
        "Regional": dados.get("Regional", ""),
        "Centro Logístico": dados.get("Centro Logístico", ""),
        "Tipo de Material": dados.get("Tipo de Material", ""),
        "Tipo de Veículo": dados.get("Tipo de Veículo", ""),
        "Qtd Veículos": dados.get("Qtd Veículos", ""),
        "Data Coleta Solicitada": dados.get("Data Coleta Solicitada", ""),
        "Data Coleta Real": "",
        "Data Entrega": dados.get("Data Entrega", ""),
        "Origem": dados.get("Origem", ""),
        "Destino": dados.get("Destino", ""),
        "Peso": dados.get("Peso", ""),
        "Quantidade": dados.get("Quantidade", ""),
        "Valor NF": dados.get("Valor NF", ""),
        "Dimensões": dados.get("Dimensões", ""),
        "Dados Complementares": dados.get("Dados Complementares", ""),
        "Arquivo": dados.get("Arquivo", ""),
        "Transportadora Vencedora": "",
        "Data Mapa Criado": "",
    }
    field_map.update(sla_map)
    return [field_map.get(header, "") for header in HEADERS_CHAMADOS]


def cotacao_to_row(chamado, transportadora):
    return [
        chamado,
        transportadora.get("empresa", ""),
        transportadora.get("contato", ""),
        transportadora.get("email", ""),
        transportadora.get("telefone", ""),
        "",
        "",
        "",
        "",
    ]


def inicializar_planilha():
    if not os.path.exists(config.PLANILHA_PATH):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Chamados"
        ws1.append(HEADERS_CHAMADOS)
        ws2 = wb.create_sheet("Cotações")
        ws2.append(HEADERS_COTACOES)
        wb.save(config.PLANILHA_PATH)
        print("✅ Planilha criada com abas Chamados e Cotações.")
    else:
        wb = load_workbook(config.PLANILHA_PATH)
        if "Cotações" not in wb.sheetnames:
            ws2 = wb.create_sheet("Cotações")
            ws2.append(HEADERS_COTACOES)
            wb.save(config.PLANILHA_PATH)
            print("✅ Aba Cotações adicionada à planilha existente.")
        wb.close()


def salvar_na_planilha(dados):
    wb = load_workbook(config.PLANILHA_PATH)
    ws = wb["Chamados"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(dados["Número Chamado"]):
            print(f"⚠️  Chamado {dados['Número Chamado']} já existe. Pulando.")
            wb.close()
            return

    agora = datetime.now()
    sla1 = dados.get("SLA Step 1")
    status = "✅ No prazo" if sla1 and agora <= sla1 else "❌ Vencido"

    ws.append(chamado_to_row(dados, status))
    wb.save(config.PLANILHA_PATH)
    wb.close()
    print(f"✅ Chamado {dados['Número Chamado']} salvo na aba Chamados.")


def salvar_cotacoes(dados, lista_transportadoras):
    wb = load_workbook(config.PLANILHA_PATH)
    ws = wb["Cotações"]
    chamado = dados.get("Número Chamado", "")

    for transportadora in lista_transportadoras:
        ws.append(cotacao_to_row(chamado, transportadora))

    wb.save(config.PLANILHA_PATH)
    wb.close()
    print(f"✅ {len(lista_transportadoras)} transportadoras salvas na aba Cotações.")


def carregar_chamado(chamado_num):
    wb = load_workbook(config.PLANILHA_PATH)
    ws = wb["Chamados"]
    headers = [c.value for c in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(chamado_num):
            return dict(zip(headers, row)), wb
    wb.close()
    return None, None


def carregar_cotacoes(chamado_num):
    wb = load_workbook(config.PLANILHA_PATH)
    ws = wb["Cotações"]
    headers = [c.value for c in ws[1]]
    cotacoes = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(chamado_num):
            cotacoes.append(dict(zip(headers, row)))
    wb.close()
    return cotacoes


def atualizar_chamado_vencedor(chamado_num, vencedor, data_mapa):
    wb = load_workbook(config.PLANILHA_PATH)
    ws = wb["Chamados"]
    headers = [c.value for c in ws[1]]

    try:
        col_vencedor = headers.index("Transportadora Vencedora") + 1
        col_data = headers.index("Data Mapa Criado") + 1
    except ValueError:
        print("⚠️  Colunas Transportadora Vencedora / Data Mapa Criado não encontradas.")
        wb.close()
        return

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(chamado_num):
            ws.cell(row=row[0].row, column=col_vencedor).value = vencedor
            ws.cell(row=row[0].row, column=col_data).value = data_mapa
            break

    wb.save(config.PLANILHA_PATH)
    wb.close()
    print(f"✅ Chamados tab atualizado — Vencedor: {vencedor} | Data: {data_mapa}")
