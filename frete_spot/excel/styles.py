"""Excel styling utilities shared by workbook formatting and map generation."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

GREEN_DARK = PatternFill("solid", fgColor="1F6B3A")
GREEN_LIGHT = PatternFill("solid", fgColor="C6EFCE")
RED_LIGHT = PatternFill("solid", fgColor="FFC7CE")
GREY_LIGHT = PatternFill("solid", fgColor="D9D9D9")
WHITE = PatternFill("solid", fgColor="FFFFFF")
WHITE_FILL = WHITE
GREY_ROW = PatternFill("solid", fgColor="F2F2F2")
YELLOW_EMPTY = PatternFill("solid", fgColor="FFFFC7")
RED_SLA = PatternFill("solid", fgColor="FFC7CE")
YELLOW_SLA = PatternFill("solid", fgColor="FFEB9C")
GREEN_SLA = PatternFill("solid", fgColor="C6EFCE")

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_NORMAL = Font(name="Calibri", bold=False, color="000000", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, color="000000", size=10)
FONT_WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_BLACK_BOLD = Font(name="Calibri", bold=True, color="000000", size=10)
FONT_BLACK = Font(name="Calibri", bold=False, color="000000", size=10)
FONT_GREEN_BOLD = Font(name="Calibri", bold=True, color="1F6B3A", size=11)
FONT_SLA_RED = Font(name="Calibri", bold=True, color="9C0006", size=10)
FONT_SLA_YELLOW = Font(name="Calibri", bold=True, color="9C6500", size=10)
FONT_SLA_GREEN = Font(name="Calibri", bold=True, color="276221", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def apply(ws, row, col, value=None, fill=None, font=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = thin_border()
    return cell


def set_column_widths(ws, headers, widths, default=16):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = widths.get(header, default)


def style_header_row(ws, headers, border=None):
    border = border or thin_border()
    ws.row_dimensions[1].height = 30
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = GREEN_DARK
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = border
