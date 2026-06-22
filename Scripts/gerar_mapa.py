import os
import re
from datetime import datetime
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

# ============================================================
# PATHS
# ============================================================
BASE          = r"C:\Users\ana.monticelli\OneDrive\OneDrive - MRV\Área de Trabalho\Frete_Spot_Monticelli"
PLANILHA_PATH = os.path.join(BASE, "program_monticelli.xlsx")
MAPAS_DIR     = os.path.join(BASE, "3_Mapas_Cotacao")

# ============================================================
# STYLES
# ============================================================
GREEN_DARK   = PatternFill("solid", fgColor="1F6B3A")   # header dark green
GREEN_LIGHT  = PatternFill("solid", fgColor="C6EFCE")   # winner
RED_LIGHT    = PatternFill("solid", fgColor="FFC7CE")   # loser
GREY_LIGHT   = PatternFill("solid", fgColor="D9D9D9")   # label column
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

FONT_WHITE_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_BLACK_BOLD  = Font(name="Calibri", bold=True, color="000000", size=10)
FONT_BLACK       = Font(name="Calibri", bold=False, color="000000", size=10)
FONT_GREEN_BOLD  = Font(name="Calibri", bold=True, color="1F6B3A", size=11)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply(ws, row, col, value=None, fill=None, font=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = thin_border()
    return cell

# ============================================================
# LOAD DATA
# ============================================================
def carregar_chamado(chamado_num):
    wb = load_workbook(PLANILHA_PATH)
    ws = wb["Chamados"]
    headers = [c.value for c in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(chamado_num):
            return dict(zip(headers, row)), wb
    wb.close()
    return None, None

def carregar_cotacoes(chamado_num):
    wb = load_workbook(PLANILHA_PATH)
    ws = wb["Cotações"]
    headers = [c.value for c in ws[1]]
    cotacoes = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(chamado_num):
            d = dict(zip(headers, row))
            cotacoes.append(d)
    wb.close()
    return cotacoes

# ============================================================
# UPDATE CHAMADOS TAB — winner + date
# ============================================================
def atualizar_chamado_vencedor(chamado_num, vencedor, data_mapa):
    wb = load_workbook(PLANILHA_PATH)
    ws = wb["Chamados"]
    headers = [c.value for c in ws[1]]

    try:
        col_vencedor = headers.index("Transportadora Vencedora") + 1
        col_data     = headers.index("Data Mapa Criado") + 1
    except ValueError:
        print("⚠️  Colunas Transportadora Vencedora / Data Mapa Criado não encontradas.")
        wb.close()
        return

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(chamado_num):
            ws.cell(row=row[0].row, column=col_vencedor).value = vencedor
            ws.cell(row=row[0].row, column=col_data).value     = data_mapa
            break

    wb.save(PLANILHA_PATH)
    wb.close()
    print(f"✅ Chamados tab atualizado — Vencedor: {vencedor} | Data: {data_mapa}")

# ============================================================
# GENERATE MAPA
# ============================================================
def gerar_mapa(chamado_num):
    os.makedirs(MAPAS_DIR, exist_ok=True)

    chamado, wb_main = carregar_chamado(chamado_num)
    if not chamado:
        print(f"❌ Chamado {chamado_num} não encontrado na planilha.")
        return
    wb_main.close()

    cotacoes = carregar_cotacoes(chamado_num)

    # Filter only responded quotes (have a price)
    respondidas = []
    for c in cotacoes:
        val = c.get("Valor Frete (R$)", "")
        if val != "" and val is not None:
            try:
                float(str(val).replace("R$", "").replace(".", "").replace(",", ".").strip())
                respondidas.append(c)
            except:
                pass

    if not respondidas:
        print("⚠️  Nenhuma cotação com valor preenchido encontrada.")
        print("    Preencha a coluna 'Valor Frete (R$)' na aba Cotações e tente novamente.")
        return

    # Sort by price — find winner
    def parse_valor(c):
        try:
            return float(str(c.get("Valor Frete (R$)", "0"))
                         .replace("R$", "").replace(".", "").replace(",", ".").strip())
        except:
            return float("inf")

    respondidas.sort(key=parse_valor)
    vencedor = respondidas[0]

    # ── Build workbook ──────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Mapa de Cotação"

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 4

    num_fornecedores = len(respondidas)
    col_start = 6  # column F = first supplier
    cols_per_supplier = 2  # VALOR UNITARIO + VALOR TOTAL

    for i in range(num_fornecedores):
        base = col_start + i * cols_per_supplier
        ws.column_dimensions[ws.cell(row=1, column=base).column_letter].width     = 14
        ws.column_dimensions[ws.cell(row=1, column=base+1).column_letter].width   = 14

    # ── Row 4 — FORNECEDOR headers ───────────────────────────
    for i, cot in enumerate(respondidas):
        base = col_start + i * cols_per_supplier
        label = f"FORNECEDOR {str(i+1).zfill(2)}"
        ws.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base+1)
        cell = ws.cell(row=4, column=base)
        cell.value     = label
        cell.fill      = GREEN_DARK
        cell.font      = FONT_WHITE_BOLD
        cell.alignment = CENTER
        cell.border    = thin_border()

    # ── Rows 5–12 — Info block ───────────────────────────────
    labels = ["CÓDIGO", "FORNECEDOR", "TELEFONE", "E-MAIL",
              "CONTATO", "PRAZO ENTREGA", "PRAZO DE PAGAMENTO", "OBSERVAÇÃO"]
    fields = ["Número Chamado", "Transportadora", "Telefone",
              "Email", "Contato", "Prazo Entrega", "_pagamento", "Observações"]

    for idx, (label, field) in enumerate(zip(labels, fields)):
        r = 5 + idx

        # Merge D+E for label
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        apply(ws, r, 4, label, GREEN_DARK, FONT_WHITE_BOLD, CENTER)

        for i, cot in enumerate(respondidas):
            base = col_start + i * cols_per_supplier
            ws.merge_cells(start_row=r, start_column=base, end_row=r, end_column=base+1)
            cell = ws.cell(row=r, column=base)

            if field == "_pagamento":
                val = "30 Dias"
            elif field == "Número Chamado":
                val = chamado_num
            else:
                val = cot.get(field, "")

            cell.value     = val
            cell.fill      = WHITE_FILL
            cell.font      = FONT_BLACK
            cell.alignment = CENTER
            cell.border    = thin_border()

    # ── Row 13 — Table headers ───────────────────────────────
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=3)
    apply(ws, 13, 2, "DESCRIÇÃO",  GREY_LIGHT, FONT_BLACK_BOLD, CENTER)
    apply(ws, 13, 4, "QTD",        GREY_LIGHT, FONT_BLACK_BOLD, CENTER)
    ws.merge_cells(start_row=13, start_column=4, end_row=13, end_column=5)

    for i in range(num_fornecedores):
        base = col_start + i * cols_per_supplier
        apply(ws, 13, base,   "VALOR\nUNITARIO", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)
        apply(ws, 13, base+1, "VALOR TOTAL",     GREY_LIGHT, FONT_BLACK_BOLD, CENTER)

    # ── Row 14 — Service line ────────────────────────────────
    assunto    = chamado.get("Assunto", "Frete")
    tipo_mat   = chamado.get("Tipo de Material", "")
    descricao  = f"{assunto}" + (f" - {tipo_mat}" if tipo_mat else "")

    ws.merge_cells(start_row=14, start_column=2, end_row=14, end_column=3)
    apply(ws, 14, 2, descricao, WHITE_FILL, FONT_BLACK, CENTER)

    ws.merge_cells(start_row=14, start_column=4, end_row=14, end_column=5)
    apply(ws, 14, 4, 1, WHITE_FILL, FONT_BLACK, CENTER)

    winner_val = parse_valor(vencedor)

    for i, cot in enumerate(respondidas):
        base  = col_start + i * cols_per_supplier
        val   = parse_valor(cot)
        is_winner = (cot == vencedor)
        fill  = GREEN_LIGHT if is_winner else RED_LIGHT

        # VALOR UNITARIO
        cell_u = ws.cell(row=14, column=base)
        cell_u.value     = val
        cell_u.number_format = 'R$ #,##0.00'
        cell_u.fill      = fill
        cell_u.font      = FONT_BLACK_BOLD
        cell_u.alignment = CENTER
        cell_u.border    = thin_border()

        # VALOR TOTAL
        cell_t = ws.cell(row=14, column=base+1)
        cell_t.value     = val
        cell_t.number_format = 'R$ #,##0.00'
        cell_t.fill      = fill
        cell_t.font      = FONT_BLACK_BOLD
        cell_t.alignment = CENTER
        cell_t.border    = thin_border()

    # ── Row 15 — VALOR TOTAL row ─────────────────────────────
    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=5)
    apply(ws, 15, 2, "VALOR TOTAL", GREY_LIGHT, FONT_BLACK_BOLD, CENTER)

    for i, cot in enumerate(respondidas):
        base     = col_start + i * cols_per_supplier
        val      = parse_valor(cot)
        is_winner = (cot == vencedor)
        fill     = GREEN_LIGHT if is_winner else RED_LIGHT

        ws.merge_cells(start_row=15, start_column=base, end_row=15, end_column=base+1)
        cell = ws.cell(row=15, column=base)
        cell.value        = val
        cell.number_format = 'R$ #,##0.00'
        cell.fill         = fill
        cell.font         = FONT_BLACK_BOLD
        cell.alignment    = CENTER
        cell.border       = thin_border()

    # ── Title block (left side rows 6–10) ────────────────────
    ws.merge_cells(start_row=6, start_column=2, end_row=10, end_column=3)
    title_cell = ws.cell(row=6, column=2)
    title_cell.value     = f"#{chamado_num}\n{assunto}"
    title_cell.font      = FONT_GREEN_BOLD
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border    = thin_border()

    # ── Save ─────────────────────────────────────────────────
    data_hoje  = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"Mapa_{chamado_num}_{data_hoje}.xlsx"
    caminho      = os.path.join(MAPAS_DIR, nome_arquivo)
    wb.save(caminho)
    print(f"\n✅ Mapa salvo em: 3_Mapas_Cotacao/{nome_arquivo}")

    # ── Update Chamados tab ───────────────────────────────────
    atualizar_chamado_vencedor(
        chamado_num,
        vencedor.get("Transportadora", ""),
        data_hoje
    )

    print(f"\n🏆 Vencedor  : {vencedor.get('Transportadora')} — R$ {parse_valor(vencedor):,.2f}")
    print(f"📅 Data Mapa : {data_hoje}")
    print(f"\n📊 Resumo das cotações:")
    for i, c in enumerate(respondidas):
        tag = " 🏆 VENCEDOR" if c == vencedor else ""
        print(f"   {i+1}. {c.get('Transportadora'):25s} R$ {parse_valor(c):>10,.2f}{tag}")

# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 50)
    print("   GERADOR DE MAPA DE COTAÇÃO")
    print("=" * 50)
    chamado_num = input("\n👉 Digite o número do chamado: ").strip()

    if not chamado_num:
        print("❌ Número do chamado não informado.")
        return

    gerar_mapa(chamado_num)

if __name__ == "__main__":
    main()
