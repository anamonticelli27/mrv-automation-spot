import os
import re
import shutil
import win32com.client
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import gc
import time

# ============================================================
# PATHS
# ============================================================
BASE          = r"C:\Users\ana.monticelli\OneDrive\OneDrive - MRV\Área de Trabalho\Frete_Spot_Monticelli"
ENTRADA       = os.path.join(BASE, "1_Agilis_Emails_Entrada")
PROCESSADOS   = os.path.join(BASE, "2_Agilis_Emails_Processados")
PLANILHA_PATH = os.path.join(BASE, "program_monticelli.xlsx")

# ============================================================
# TRANSPORTADORAS
# ============================================================
FRACIONADO = [
    {"empresa": "GMS Transportes",  "contato": "Geraldo Leite",     "email": "geraldoleite@dedetransportes.com.br", "telefone": ""},
    {"empresa": "DEDE",             "contato": "Vyctor Carvalho",   "email": "vyctor@dedetransportes.com.br",       "telefone": ""},
    {"empresa": "IRST",             "contato": "Climério Neto",     "email": "neto.ll@hotmail.com",                 "telefone": ""},
    {"empresa": "BRIX",             "contato": "Raiane Fernandes",  "email": "raiane.fernandes@brixcargo.com.br",   "telefone": "31 98457 4885"},
    {"empresa": "HT LOG",           "contato": "Lucas Nicoladelli", "email": "lucas@htlog.com.br",                  "telefone": ""},
    {"empresa": "VIAMAP",           "contato": "Jeison Mateus",     "email": "comercial@viamap.com.br",             "telefone": ""},
    {"empresa": "TH TRANSPORTES",   "contato": "Mauro Alves",       "email": "thtransportesltda01@gmail.com",       "telefone": ""},
    {"empresa": "COZAPI",           "contato": "Darlan Rocha",      "email": "operacional@cozapi.com.br",           "telefone": ""},
]

VEICULO_COMPLETO = [
    {"empresa": "GMS Transportes",  "contato": "Geraldo Leite",     "email": "geraldoleite@dedetransportes.com.br", "telefone": ""},
    {"empresa": "DEDE",             "contato": "Vyctor Carvalho",   "email": "vyctor@dedetransportes.com.br",       "telefone": ""},
    {"empresa": "HT LOG",           "contato": "Lucas Nicoladelli", "email": "lucas@htlog.com.br",                  "telefone": ""},
    {"empresa": "VIAMAP",           "contato": "Jeison Mateus",     "email": "comercial@viamap.com.br",             "telefone": ""},
    {"empresa": "COZAPI",           "contato": "Darlan Rocha",      "email": "operacional@cozapi.com.br",           "telefone": ""},
]

MUDANCA = [
    {"empresa": "Montreal Mudanças",          "contato": "Valter Araujo", "email": "comercial@montrealmudancas.com.br",          "telefone": ""},
    {"empresa": "Fibra Mudanças",             "contato": "",              "email": "Contato@fibramudancas.com.br",               "telefone": ""},
    {"empresa": "Fibra Mudanças",             "contato": "Gentil Neto",   "email": "gentil.neto@fibramudancas.com.br",           "telefone": ""},
    {"empresa": "Moriway",                    "contato": "",              "email": "comercial@moriway.com.br",                   "telefone": ""},
    {"empresa": "CONFIANÇA MUDANÇA",          "contato": "",              "email": "conf.mcz@confiancabr.com.br",                "telefone": ""},
    {"empresa": "CONFIANÇA MUDANÇA",          "contato": "",              "email": "bocaliniwilson@yahoo.com.br",                "telefone": ""},
    {"empresa": "Mudanças Tucuruvi",          "contato": "Tatiana",       "email": "tatiana@mudancastucuruvi.com.br",            "telefone": ""},
    {"empresa": "Mudanças Trans Continental", "contato": "",              "email": "mudancas@mudancastranscontinental.com.br",   "telefone": ""},
    {"empresa": "Lord Mudanças",              "contato": "",              "email": "vendas02@lordmudancas.com.br",               "telefone": ""},
    {"empresa": "Personnalite Transportes",   "contato": "",              "email": "comercial01@personnalitetransportes.com.br", "telefone": ""},
]

# ============================================================
# CATEGORY DETECTION
# ============================================================
def determinar_categoria(subcategoria):
    sub = subcategoria.lower().strip()
    if "pessoas" in sub:
        return "PESSOAS"
    elif "expresso" in sub or "expres" in sub:
        return "EXPRESS_FRETE"
    elif "mudan" in sub:
        return "MUDANCA"
    else:
        return "NORMAL_FRETE"

def determinar_lista_transportadoras(categoria, tipo_veiculo):
    vei = tipo_veiculo.lower().strip()
    if categoria == "MUDANCA":
        return MUDANCA
    elif "fracionado" in vei:
        return FRACIONADO
    else:
        return VEICULO_COMPLETO

# ============================================================
# BUSINESS DAYS CALCULATOR
# ============================================================
def adicionar_dias_uteis(data_inicio, dias):
    data = data_inicio
    dias_adicionados = 0
    while dias_adicionados < dias:
        data += timedelta(days=1)
        if data.weekday() < 5:
            dias_adicionados += 1
    return data

# ============================================================
# SLA CALCULATION
# ============================================================
SLA_RULES = {
    "NORMAL_FRETE":  (2, 6, 0, 4),
    "EXPRESS_FRETE": (1, 2, 0, 1),
    "MUDANCA":       (2, 5, 1, 2),
    "PESSOAS":       (2, 2, 0, 1),
}

def calcular_slas(dt_recebimento, categoria):
    regras = SLA_RULES.get(categoria, (2, 6, 0, 4))
    step1_days, step2_days, step3_days, step4_days = regras

    sla_step1 = adicionar_dias_uteis(dt_recebimento, step1_days)
    sla_step2 = adicionar_dias_uteis(sla_step1, step2_days)

    if step3_days > 0:
        sla_step3 = adicionar_dias_uteis(sla_step2, step3_days)
        sla_step4 = adicionar_dias_uteis(sla_step3, step4_days)
    else:
        sla_step3 = None
        sla_step4 = adicionar_dias_uteis(sla_step2, step4_days)

    return sla_step1, sla_step2, sla_step3, sla_step4

# ============================================================
# FIELD EXTRACTION
# ============================================================
def extrair_campo(body, campo):
    pattern = rf"\*\s*{re.escape(campo)}\s*:\s*(.+)"
    match = re.search(pattern, body, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return ""

def extrair_numero_chamado(subject, body):
    match = re.search(r"##RE-(\d+)##", subject)
    if match:
        return match.group(1)
    match = re.search(r"N[uú]mero do Chamd?o\s*:\s*(\d+)", body, re.IGNORECASE)
    if match:
        return match.group(1)
    return ""

def extrair_assunto(body):
    match = re.search(r"\*\s*Assunto\s*:\s*(.+)", body, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return ""

def extrair_pep(body):
    match = re.search(
        r"\*\s*N[uú]mero do Elemento PEP.*?:\s*(.+)",
        body, re.IGNORECASE
    )
    if match:
        return match.group(1).strip()
    return ""

def extrair_codigo_centro_logistico(body):
    raw = extrair_campo(body, "Centro Logístico")
    if raw:
        match = re.search(r"\b([A-Z]{2}\d{2})\b", raw.strip())
        if match:
            return match.group(1)
    return raw

# ============================================================
# SPREADSHEET — HEADERS
# ============================================================
HEADERS_CHAMADOS = [
    "Número Chamado", "Assunto", "PEP / Diagrama de Rede",
    "Data Recebimento", "Categoria",
    "SLA Step 1 (Validar)", "SLA Step 2 (Cotar)",
    "SLA Step 3 (OK Cliente)", "SLA Step 4 (Programar)",
    "Status Geral",
    "Subcategoria", "Solicitante", "Email Solicitante",
    "Telefone", "Regional", "Centro Logístico",
    "Tipo de Material", "Tipo de Veículo", "Qtd Veículos",
    "Data Coleta Solicitada", "Data Coleta Real",
    "Data Entrega",
    "Origem", "Destino",
    "Peso", "Quantidade", "Valor NF",
    "Dimensões", "Dados Complementares", "Arquivo",
    "Transportadora Vencedora", "Data Mapa Criado"
]

HEADERS_COTACOES = [
    "Número Chamado", "Transportadora", "Contato",
    "Email", "Telefone",
    "Data Resposta", "Valor Frete (R$)", "Prazo Entrega", "Observações"
]

def fmt_data(dt):
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y")
    return ""

# ============================================================
# SPREADSHEET — INIT
# ============================================================
def inicializar_planilha():
    if not os.path.exists(PLANILHA_PATH):
        wb = openpyxl.Workbook()

        # Tab 1 — Chamados
        ws1 = wb.active
        ws1.title = "Chamados"
        ws1.append(HEADERS_CHAMADOS)

        # Tab 2 — Cotações
        ws2 = wb.create_sheet("Cotações")
        ws2.append(HEADERS_COTACOES)

        wb.save(PLANILHA_PATH)
        print("✅ Planilha criada com abas Chamados e Cotações.")
    else:
        wb = load_workbook(PLANILHA_PATH)
        if "Cotações" not in wb.sheetnames:
            ws2 = wb.create_sheet("Cotações")
            ws2.append(HEADERS_COTACOES)
            wb.save(PLANILHA_PATH)
            print("✅ Aba Cotações adicionada à planilha existente.")
        wb.close()

# ============================================================
# SPREADSHEET — SAVE CHAMADO
# ============================================================
def salvar_na_planilha(dados):
    wb = load_workbook(PLANILHA_PATH)
    ws = wb["Chamados"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(dados["Número Chamado"]):
            print(f"⚠️  Chamado {dados['Número Chamado']} já existe. Pulando.")
            wb.close()
            return

    agora  = datetime.now()
    sla1   = dados.get("SLA Step 1")
    status = "✅ No prazo" if sla1 and agora <= sla1 else "❌ Vencido"

    ws.append([
        dados.get("Número Chamado", ""),
        dados.get("Assunto", ""),
        dados.get("PEP", ""),
        fmt_data(dados.get("Data Recebimento")),
        dados.get("Categoria", ""),
        fmt_data(dados.get("SLA Step 1")),
        fmt_data(dados.get("SLA Step 2")),
        fmt_data(dados.get("SLA Step 3")),
        fmt_data(dados.get("SLA Step 4")),
        status,
        dados.get("Subcategoria", ""),
        dados.get("Solicitante", ""),
        dados.get("Email Solicitante", ""),
        dados.get("Telefone", ""),
        dados.get("Regional", ""),
        dados.get("Centro Logístico", ""),
        dados.get("Tipo de Material", ""),
        dados.get("Tipo de Veículo", ""),
        dados.get("Qtd Veículos", ""),
        dados.get("Data Coleta Solicitada", ""),
        "",  # Data Coleta Real — manual
        dados.get("Data Entrega", ""),
        dados.get("Origem", ""),
        dados.get("Destino", ""),
        dados.get("Peso", ""),
        dados.get("Quantidade", ""),
        dados.get("Valor NF", ""),
        dados.get("Dimensões", ""),
        dados.get("Dados Complementares", ""),
        dados.get("Arquivo", ""),
        "",  # Transportadora Vencedora — filled by gerar_mapa
        "",  # Data Mapa Criado — filled by gerar_mapa
    ])
    wb.save(PLANILHA_PATH)
    print(f"✅ Chamado {dados['Número Chamado']} salvo na aba Chamados.")

# ============================================================
# SPREADSHEET — SAVE COTAÇÕES
# ============================================================
def salvar_cotacoes(dados, lista_transportadoras):
    wb = load_workbook(PLANILHA_PATH)
    ws = wb["Cotações"]

    chamado = dados.get("Número Chamado", "")

    for t in lista_transportadoras:
        ws.append([
            chamado,
            t.get("empresa", ""),
            t.get("contato", ""),
            t.get("email", ""),
            t.get("telefone", ""),
            "",  # Data Resposta — manual
            "",  # Valor Frete — manual
            "",  # Prazo Entrega — manual
            "",  # Observações — manual
        ])

    wb.save(PLANILHA_PATH)
    print(f"✅ {len(lista_transportadoras)} transportadoras salvas na aba Cotações.")

# ============================================================
# EMAIL DRAFT BODIES
# ============================================================
def gerar_corpo_frete(dados, modalidade):
    return f"""Prezado Transportador,

Boa tarde!

Solicito cotação de frete conforme dados abaixo:

• Chamado Agilis : {dados.get('Número Chamado', '')}
• Modalidade     : {modalidade}
• Tipo Material  : {dados.get('Tipo de Material', '')}
• Origem         : {dados.get('Origem', '')}
• Destino        : {dados.get('Destino', '')}
• Data Coleta    : {dados.get('Data Coleta Solicitada', '')}
• Data Entrega   : {dados.get('Data Entrega', '')}
• Peso           : {dados.get('Peso', '')}
• Quantidade     : {dados.get('Quantidade', '')}
• Dimensões      : {dados.get('Dimensões', '')}
• Valor da NF    : {dados.get('Valor NF', '')}
• Observações    : {dados.get('Dados Complementares', '')}

Por favor, retornar com:
✔ Valor do frete
✔ Prazo de entrega
✔ Código interno (se houver)

Aguardo retorno.

Atenciosamente,
Ana Laura Monticelli
MRV Engenharia"""

def gerar_corpo_mudanca(dados):
    return f"""Prezado Transportador,

Boa tarde!

Solicito cotação de mudança conforme dados abaixo:

• Chamado Agilis : {dados.get('Número Chamado', '')}
• Solicitante    : {dados.get('Solicitante', '')}
• Origem         : {dados.get('Origem', '')}
• Destino        : {dados.get('Destino', '')}
• Data Coleta    : {dados.get('Data Coleta Solicitada', '')}
• Data Entrega   : {dados.get('Data Entrega', '')}
• Observações    : {dados.get('Dados Complementares', '')}

Por favor, retornar com:
✔ Valor do frete
✔ Prazo de entrega
✔ Código interno (se houver)

Aguardo retorno.

Atenciosamente,
Ana Laura Monticelli
MRV Engenharia"""

# ============================================================
# CREATE OUTLOOK DRAFT
# ============================================================
def criar_rascunho_outlook(dados, categoria):
    chamado = dados.get("Número Chamado", "SEM_NUMERO")
    lista   = determinar_lista_transportadoras(categoria, dados.get("Tipo de Veículo", ""))

    if categoria == "MUDANCA":
        corpo = gerar_corpo_mudanca(dados)
    elif categoria == "EXPRESS_FRETE":
        corpo = gerar_corpo_frete(dados, "Frete Expresso")
    elif categoria == "PESSOAS":
        corpo = gerar_corpo_frete(dados, "Frete Expresso - Pessoas")
    else:
        corpo = gerar_corpo_frete(dados, dados.get("Tipo de Veículo", "Frete Normal"))

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail    = outlook.CreateItem(0)

        mail.Subject = f"COTAÇÃO SERVIÇO DE TRANSPORTE DE MATERIAL - {chamado}"
        mail.Body    = corpo
        mail.To      = ""

        for t in lista:
            recipient      = mail.Recipients.Add(t["email"])
            recipient.Type = 3  # olBCC

        mail.Recipients.ResolveAll()
        mail.Save()

        print(f"   📧 Rascunho criado no Outlook para chamado {chamado} ({len(lista)} transportadoras em BCC)")

    except Exception as e:
        print(f"   ❌ Erro ao criar rascunho no Outlook: {e}")

# ============================================================
# PROCESS .MSG
# ============================================================
def processar_msg(filepath):
    outlook_ns = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    msg        = outlook_ns.OpenSharedItem(filepath)

    subject  = msg.Subject or ""
    body     = msg.Body or ""
    received = msg.ReceivedTime

    try:
        dt_recebimento = datetime(
            received.year, received.month, received.day,
            received.hour, received.minute, received.second
        )
    except:
        dt_recebimento = datetime.now()

    subcategoria = extrair_campo(body, "Subcategoria")
    categoria    = determinar_categoria(subcategoria)
    sla1, sla2, sla3, sla4 = calcular_slas(dt_recebimento, categoria)

    dados = {
        "Número Chamado":         extrair_numero_chamado(subject, body),
        "Assunto":                extrair_assunto(body),
        "PEP":                    extrair_pep(body),
        "Data Recebimento":       dt_recebimento,
        "Categoria":              categoria,
        "SLA Step 1":             sla1,
        "SLA Step 2":             sla2,
        "SLA Step 3":             sla3,
        "SLA Step 4":             sla4,
        "Subcategoria":           subcategoria,
        "Solicitante":            extrair_campo(body, "Solicitante"),
        "Email Solicitante":      extrair_campo(body, "E-mail"),
        "Telefone":               extrair_campo(body, "Telefone/Contato"),
        "Regional":               extrair_campo(body, "Regional Suprimentos"),
        "Centro Logístico":       extrair_codigo_centro_logistico(body),
        "Tipo de Material":       extrair_campo(body, "Tipo de Material"),
        "Tipo de Veículo":        extrair_campo(body, "Tipo de veículo"),
        "Qtd Veículos":           extrair_campo(body, "Quantidade de veículos necessários para carregamento"),
        "Data Coleta Solicitada": extrair_campo(body, "Data da Coleta na Origem"),
        "Data Entrega":           extrair_campo(body, "Data de Entrega no Destino"),
        "Origem":                 extrair_campo(body, "Endereço de Origem - Completo"),
        "Destino":                extrair_campo(body, "Endereço de Destino - Completo"),
        "Peso":                   extrair_campo(body, "Peso"),
        "Quantidade":             extrair_campo(body, "Quantidade"),
        "Valor NF":               extrair_campo(body, "Valor da NF"),
        "Dimensões":              extrair_campo(body, "Dimensões"),
        "Dados Complementares":   extrair_campo(body, "Dados Complementares"),
        "Arquivo":                os.path.basename(filepath),
    }

    print(f"\n📨 Chamado  : {dados['Número Chamado']}")
    print(f"   Assunto   : {dados['Assunto']}")
    print(f"   PEP       : {dados['PEP']}")
    print(f"   Categoria : {categoria}")
    print(f"   SLA Step1 : {fmt_data(sla1)}")
    print(f"   SLA Step2 : {fmt_data(sla2)}")
    if sla3:
        print(f"   SLA Step3 : {fmt_data(sla3)}")
    print(f"   SLA Step4 : {fmt_data(sla4)}")

    lista_transportadoras = determinar_lista_transportadoras(categoria, dados.get("Tipo de Veículo", ""))

    salvar_na_planilha(dados)
    salvar_cotacoes(dados, lista_transportadoras)
    criar_rascunho_outlook(dados, categoria)

    msg.Close(0)
    del msg
    gc.collect()
    time.sleep(1)

    destino = os.path.join(PROCESSADOS, os.path.basename(filepath))
    try:
        shutil.move(filepath, destino)
        print(f"   📁 Movido para: 2_Agilis_Emails_Processados")
    except Exception as e:
        print(f"   ⚠️  Não foi possível mover o arquivo: {e}")

# ============================================================
# AUTO-FORMAT SPREADSHEET
# ============================================================
def formatar_planilha():
    wb = load_workbook(PLANILHA_PATH)

    # ── COLORS ──────────────────────────────────────────────
    GREEN_DARK   = PatternFill("solid", fgColor="1F6B3A")
    WHITE        = PatternFill("solid", fgColor="FFFFFF")
    GREY_ROW     = PatternFill("solid", fgColor="F2F2F2")
    YELLOW_EMPTY = PatternFill("solid", fgColor="FFFFC7")
    RED_SLA      = PatternFill("solid", fgColor="FFC7CE")
    YELLOW_SLA   = PatternFill("solid", fgColor="FFEB9C")
    GREEN_SLA    = PatternFill("solid", fgColor="C6EFCE")

    # ── FONTS ───────────────────────────────────────────────
    FONT_HEADER  = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
    FONT_NORMAL  = Font(name="Calibri", bold=False, color="000000", size=10)
    FONT_BOLD    = Font(name="Calibri", bold=True,  color="000000", size=10)

    # ── ALIGNMENT ───────────────────────────────────────────
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── BORDER ──────────────────────────────────────────────
    thin   = Side(style="thin")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    today = datetime.now().date()

    # ════════════════════════════════════════════════════════
    # TAB 1 — CHAMADOS
    # ════════════════════════════════════════════════════════
    ws1 = wb["Chamados"]
    headers1 = [c.value for c in ws1[1]]

    col_widths_chamados = {
        "Número Chamado":              16,
        "Assunto":                     40,
        "PEP / Diagrama de Rede":      22,
        "Data Recebimento":            18,
        "Categoria":                   18,
        "SLA Step 1 (Validar)":        18,
        "SLA Step 2 (Cotar)":          18,
        "SLA Step 3 (OK Cliente)":     20,
        "SLA Step 4 (Programar)":      20,
        "Status Geral":                14,
        "Subcategoria":                22,
        "Solicitante":                 22,
        "Email Solicitante":           30,
        "Telefone":                    16,
        "Regional":                    16,
        "Centro Logístico":            16,
        "Tipo de Material":            22,
        "Tipo de Veículo":             18,
        "Qtd Veículos":                14,
        "Data Coleta Solicitada":      20,
        "Data Coleta Real":            18,
        "Data Entrega":                16,
        "Origem":                      35,
        "Destino":                     35,
        "Peso":                        12,
        "Quantidade":                  12,
        "Valor NF":                    14,
        "Dimensões":                   18,
        "Dados Complementares":        30,
        "Arquivo":                     30,
        "Transportadora Vencedora":    24,
        "Data Mapa Criado":            18,
    }

    for col_idx, header in enumerate(headers1, start=1):
        col_letter = ws1.cell(row=1, column=col_idx).column_letter
        ws1.column_dimensions[col_letter].width = col_widths_chamados.get(header, 16)

    sla_cols = []
    for name in ["SLA Step 1 (Validar)", "SLA Step 2 (Cotar)",
                 "SLA Step 3 (OK Cliente)", "SLA Step 4 (Programar)"]:
        if name in headers1:
            sla_cols.append(headers1.index(name) + 1)

    status_col = headers1.index("Status Geral") + 1 if "Status Geral" in headers1 else None

    # Header row
    ws1.row_dimensions[1].height = 30
    for col_idx, _ in enumerate(headers1, start=1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill      = GREEN_DARK
        cell.font      = FONT_HEADER
        cell.alignment = CENTER
        cell.border    = BORDER

    ws1.freeze_panes = "A2"

    # Data rows
    for row_idx in range(2, ws1.max_row + 1):
        row_fill = WHITE if row_idx % 2 == 0 else GREY_ROW
        ws1.row_dimensions[row_idx].height = 20

        for col_idx, header in enumerate(headers1, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx)

            cell.fill      = row_fill
            cell.font      = FONT_NORMAL
            cell.alignment = LEFT if header in ("Assunto", "Origem", "Destino") else CENTER
            cell.border    = BORDER

            # SLA color coding
            if col_idx in sla_cols and cell.value:
                try:
                    sla_date = datetime.strptime(str(cell.value), "%d/%m/%Y").date()
                    diff = (sla_date - today).days
                    if diff < 0:
                        cell.fill = RED_SLA
                        cell.font = Font(name="Calibri", bold=True, color="9C0006", size=10)
                    elif diff <= 2:
                        cell.fill = YELLOW_SLA
                        cell.font = Font(name="Calibri", bold=True, color="9C6500", size=10)
                    else:
                        cell.fill = GREEN_SLA
                        cell.font = Font(name="Calibri", bold=True, color="276221", size=10)
                except:
                    pass

            # Status column
            if status_col and col_idx == status_col:
                val = str(cell.value or "")
                if "Vencido" in val:
                    cell.fill = RED_SLA
                    cell.font = Font(name="Calibri", bold=True, color="9C0006", size=10)
                elif "prazo" in val.lower():
                    cell.fill = GREEN_SLA
                    cell.font = Font(name="Calibri", bold=True, color="276221", size=10)

            # Transportadora Vencedora — bold
            if header == "Transportadora Vencedora" and cell.value:
                cell.font = FONT_BOLD

    # ════════════════════════════════════════════════════════
    # TAB 2 — COTAÇÕES
    # ════════════════════════════════════════════════════════
    ws2 = wb["Cotações"]
    headers2 = [c.value for c in ws2[1]]

    col_widths_cotacoes = {
        "Número Chamado":    16,
        "Transportadora":    24,
        "Contato":           22,
        "Email":             32,
        "Telefone":          16,
        "Data Resposta":     16,
        "Valor Frete (R$)":  16,
        "Prazo Entrega":     16,
        "Observações":       35,
    }

    for col_idx, header in enumerate(headers2, start=1):
        col_letter = ws2.cell(row=1, column=col_idx).column_letter
        ws2.column_dimensions[col_letter].width = col_widths_cotacoes.get(header, 16)

    valor_col = headers2.index("Valor Frete (R$)") + 1 if "Valor Frete (R$)" in headers2 else None
    data_col  = headers2.index("Data Resposta")    + 1 if "Data Resposta"    in headers2 else None

    # Header row
    ws2.row_dimensions[1].height = 30
    for col_idx, _ in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill      = GREEN_DARK
        cell.font      = FONT_HEADER
        cell.alignment = CENTER
        cell.border    = BORDER

    ws2.freeze_panes = "A2"

    # Data rows
    for row_idx in range(2, ws2.max_row + 1):
        row_fill = WHITE if row_idx % 2 == 0 else GREY_ROW
        ws2.row_dimensions[row_idx].height = 20

        data_val  = ws2.cell(row=row_idx, column=data_col).value  if data_col  else None
        valor_val = ws2.cell(row=row_idx, column=valor_col).value if valor_col else None
        sem_resposta = (not data_val and not valor_val)

        for col_idx, header in enumerate(headers2, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx)

            cell.fill      = YELLOW_EMPTY if sem_resposta else row_fill
            cell.font      = FONT_NORMAL
            cell.alignment = LEFT if header in ("Observações", "Email") else CENTER
            cell.border    = BORDER

            # Valor Frete — bold + currency
            if col_idx == valor_col and cell.value:
                cell.font          = FONT_BOLD
                cell.number_format = 'R$ #,##0.00'

    wb.save(PLANILHA_PATH)
    print("🎨 Planilha formatada com sucesso!")

# ============================================================
# MAIN
# ============================================================
def main():
    inicializar_planilha()
    arquivos = [f for f in os.listdir(ENTRADA) if f.lower().endswith(".msg")]

    if not arquivos:
        print("📭 Nenhum email em 1_Agilis_Emails_Entrada.")
        formatar_planilha()
        return

    print(f"📬 {len(arquivos)} email(s) encontrado(s).\n")
    for arquivo in arquivos:
        filepath = os.path.join(ENTRADA, arquivo)
        try:
            processar_msg(filepath)
        except Exception as e:
            print(f"❌ Erro ao processar {arquivo}: {e}")

    formatar_planilha()
    print("\n✅ Concluído!")

if __name__ == "__main__":
    main()
